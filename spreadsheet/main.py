import openpyxl

sheet_file = openpyxl.load_workbook("inventory.xlsx")
product_list = sheet_file["Sheet1"]

product_per_supplier = {}
total_values_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation number of products per supplier
    if supplier_name in product_per_supplier:
        product_occ_count = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = product_occ_count + 1
    else:
        product_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_values_per_supplier:
        current_total_value = total_values_per_supplier.get(supplier_name)
        total_values_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_values_per_supplier[supplier_name] = inventory * price

    # logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price


print(product_per_supplier)
print(total_values_per_supplier)
print(products_under_10_inv)

sheet_file.save("inventory_with_total_value.xlsx")




